import openpyxl
from pathlib import Path
from datetime import datetime
from PyQt6.QtWidgets import QWidget, QVBoxLayout, QPushButton, QTextEdit, QLabel, QMessageBox
from openpyxl.styles import PatternFill
from FileInputRow import FileInputRow


def sync_excel_incremental(
    source_path: str,
    target_path: str,
    sheets_to_sync: list,
    key_column_name: str,
    log_path: str = None,
    mark_deleted: bool = True,
    log_func=print
):
    import openpyxl
    from openpyxl.styles import PatternFill
    from pathlib import Path
    from datetime import datetime


    source_path = Path(source_path)
    target_path = Path(target_path)

    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    log_func(f"\n[{timestamp}] 开始同步 {source_path.name} → {target_path.name}")

    log_lines = [f"\n[{timestamp}] 同步日志开始（主键列: {key_column_name}）\n"]

    source_wb = openpyxl.load_workbook(source_path)
    if not target_path.exists():
        source_wb.save(target_path)
        log_func(f"{target_path} 不存在，已创建副本。")
        return

    target_wb = openpyxl.load_workbook(target_path)

    yellow_fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
    red_fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")

    for sheet_name in sheets_to_sync:
        if sheet_name not in source_wb.sheetnames:
            log_func(f"⚠️ 源文件中没有工作表 {sheet_name}，跳过。")
            continue

        source_ws = source_wb[sheet_name]
        target_ws = target_wb[sheet_name] if sheet_name in target_wb.sheetnames else target_wb.create_sheet(sheet_name)

        # 查找主键列
        header = [cell.value for cell in source_ws[1]]
        if key_column_name not in header:
            log_func(f"❌ 源文件中未找到主键列 '{key_column_name}'，跳过 {sheet_name}")
            continue
        key_col_index = header.index(key_column_name) + 1

        # ✅ 源数据：允许重复 key
        source_data = {}
        for r in range(2, source_ws.max_row + 1):
            key = source_ws.cell(r, key_col_index).value
            if key:
                row_values = [source_ws.cell(r, c).value for c in range(1, source_ws.max_column + 1)]
                source_data.setdefault(key, []).append(row_values)

        # 目标表头
        target_header = [cell.value for cell in target_ws[1]]
        if key_column_name not in target_header:
            for c, val in enumerate(header, start=1):
                target_ws.cell(1, c).value = val
            target_header = header

        target_key_col_index = target_header.index(key_column_name) + 1

        # ✅ 目标数据映射（允许重复）
        def build_target_map(ws):
            tmap = {}
            for r in range(2, ws.max_row + 1):
                key = ws.cell(r, target_key_col_index).value
                if key:
                    tmap.setdefault(key, []).append(r)
            return tmap

        target_data = build_target_map(target_ws)

        # ✅ 删除逻辑：只删除“源中完全没有”的 key
        to_delete_keys = [k for k in target_data if k not in source_data]

        if to_delete_keys:
            if mark_deleted:
                for key in to_delete_keys:
                    for row_index in target_data[key]:
                        for c in range(1, target_ws.max_column + 1):
                            target_ws.cell(row_index, c).fill = red_fill
                    log_lines.append(f"🟥 标记删除运单号: {key} ({len(target_data[key])} 行)\n")
                    log_func(f"🟥 [{sheet_name}] 标红删除运单号: {key}")
            else:
                rows_to_delete = sorted([r for k in to_delete_keys for r in target_data[k]], reverse=True)
                for row_index in rows_to_delete:
                    target_ws.delete_rows(row_index)
                log_lines.append(f"❌ 删除多余运单号 {len(to_delete_keys)} 项。\n")
                log_func(f"❌ [{sheet_name}] 删除 {len(to_delete_keys)} 项")

        # ✅ 更新与新增逻辑
        updated, added = 0, 0
        target_data = build_target_map(target_ws)

        for key, rows in source_data.items():
            if key in target_data:
                # 更新对应的重复行（按数量对齐）
                for i, row_index in enumerate(target_data[key]):
                    if i < len(rows):
                        for c, val in enumerate(rows[i], start=1):
                            cell = target_ws.cell(row_index, c)
                            if cell.value != val:
                                old_val = cell.value
                                cell.value = val
                                cell.fill = yellow_fill
                                updated += 1
                                log_lines.append(
                                    f"🟨 [{sheet_name}] 更新 {key}: 行{row_index} 第{c}列 {old_val} → {val}\n"
                                )
                # 如果源比目标多 → 新增多余的部分
                if len(rows) > len(target_data[key]):
                    for extra in rows[len(target_data[key]):]:
                        new_row = target_ws.max_row + 1
                        for c, val in enumerate(extra, start=1):
                            target_ws.cell(new_row, c).value = val
                        added += 1
                        log_lines.append(f"➕ [{sheet_name}] 新增重复运单号 {key} (行 {new_row})\n")
            else:
                # 源中有而目标没有 → 新增
                for new in rows:
                    new_row = target_ws.max_row + 1
                    for c, val in enumerate(new, start=1):
                        target_ws.cell(new_row, c).value = val
                    added += 1
                    log_lines.append(f"➕ [{sheet_name}] 新增运单号: {key} (行 {new_row})\n")
                    log_func(f"➕ [{sheet_name}] 新增运单号: {key} (行 {new_row})")

        msg = f"✅ [{sheet_name}] 同步完成: 更新 {updated} 项, 新增 {added} 项, 删除 {len(to_delete_keys)} 运单号。\n"
        log_func(msg)
        log_lines.append(msg)

    # ✅ 保存为新文件（源文件名 + “_最新跟踪.xlsx”）
    new_target_path = target_path.parent / f"{source_path.stem}_最新跟踪.xlsx"
    target_wb.save(new_target_path)
    log_func(f"💾 已保存同步结果：{new_target_path}")

    if log_path:
        with open(log_path, "a", encoding="utf-8") as f:
            f.writelines(log_lines)
        log_func(f"🪵 日志写入：{log_path}")





# -----------------------------------
# PyQt6 界面部分



#线程更新 UI
from PyQt6.QtCore import QThread, pyqtSignal

class SyncThread(QThread):
    log_signal = pyqtSignal(str)
    done_signal = pyqtSignal(str)

    def __init__(self, f1, f2, outdir):
        super().__init__()
        self.f1, self.f2, self.outdir = f1, f2, outdir

    def run(self):
        log_file = Path(self.outdir) / "同步日志.txt"
        sync_excel_incremental(
            self.f1, self.f2,
            sheets_to_sync=["11月每日跟踪"],
            key_column_name="运单号",
            log_path=str(log_file),
            log_func=lambda msg: self.log_signal.emit(msg)
        )
        self.done_signal.emit(str(log_file))











class sync_excel_incremental_ui(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Excel智能同步工具")

        layout = QVBoxLayout()
        self.f1 = FileInputRow("最新Excel文件")
        self.f2 = FileInputRow("上一次的Excel")
        self.output_dir = FileInputRow("日志输出文件夹", folder_mode=True)

        self.run_btn = QPushButton("开始同步")
        self.run_btn.clicked.connect(self.run_organize)

        self.log = QTextEdit()
        self.log.setReadOnly(True)

        layout.addWidget(self.f1)
        layout.addWidget(self.f2)
        layout.addWidget(self.output_dir)
        layout.addWidget(self.run_btn)
        layout.addWidget(QLabel("日志输出："))
        layout.addWidget(self.log)
        self.setLayout(layout)

    def log_msg(self, msg):
        self.log.append(msg)

    def run_organize(self):
        f1, f2, outdir = self.f1.text(), self.f2.text(), self.output_dir.text()
        if not all([f1, f2, outdir]):
            QMessageBox.warning(self, "缺少文件", "请确保选择了最新Excel、上一次Excel和日志输出目录。")
            return

        try:
            self.thread = SyncThread(f1, f2, outdir)
            self.thread.log_signal.connect(self.log_msg)
            self.thread.done_signal.connect(lambda f: QMessageBox.information(self, "完成", f"同步完成：{f}"))
            self.thread.start()
        except Exception as e:
            QMessageBox.critical(self, "错误", str(e))
            self.log_msg(str(e))
