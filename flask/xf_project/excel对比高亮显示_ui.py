import os

import pandas as pd
from PyQt6.QtWidgets import QVBoxLayout, QWidget, QPushButton, QTextEdit, QLabel, QMessageBox
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from FileInputRow import FileInputRow


def highlight_differences_allow_duplicates(file1, file2, sheet_name="11月每日跟踪",
                                           key_column="运单号", output_file="diff_highlight.xlsx"):
    df1 = pd.read_excel(file1, sheet_name=sheet_name)
    df2 = pd.read_excel(file2, sheet_name=sheet_name)

    df1[key_column] = df1[key_column].astype(str).str.strip()
    df2[key_column] = df2[key_column].astype(str).str.strip()

    common_cols = df1.columns.intersection(df2.columns)
    df1 = df1[common_cols]
    df2 = df2[common_cols]

    df1_reset = df1.reset_index(drop=True)
    df1_reset.to_excel(output_file, index=False, sheet_name=sheet_name)

    wb = load_workbook(output_file)
    ws = wb[sheet_name]
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    for row_idx in range(len(df1_reset)):
        row_a = df1_reset.iloc[row_idx]
        key_value = row_a[key_column]
        df2_rows = df2[df2[key_column] == key_value]

        if df2_rows.empty:
            for col_idx in range(1, len(common_cols) + 1):
                ws.cell(row=row_idx + 2, column=col_idx).fill = yellow_fill
        else:
            row_b = df2_rows.iloc[0]
            for col_idx, col_name in enumerate(common_cols, start=1):
                val_a = row_a[col_name]
                val_b = row_b[col_name]
                str_a = "" if pd.isna(val_a) else str(val_a)
                str_b = "" if pd.isna(val_b) else str(val_b)
                if str_a != str_b:
                    ws.cell(row=row_idx + 2, column=col_idx).fill = yellow_fill

    wb.save(output_file)
    return output_file


class ExcelCompareTab(QWidget):
    def __init__(self):
        super().__init__()
        layout = QVBoxLayout()
        self.file1 = FileInputRow("今日文件：")
        self.file2 = FileInputRow("昨日文件：")
        self.output_dir = FileInputRow("输出文件夹：", folder_mode=True)
        self.run_btn = QPushButton("开始对比")
        self.run_btn.clicked.connect(self.run_compare)
        self.log = QTextEdit()
        self.log.setReadOnly(True)
        layout.addWidget(self.file1)
        layout.addWidget(self.file2)
        layout.addWidget(self.output_dir)
        layout.addWidget(self.run_btn)
        layout.addWidget(QLabel("日志输出："))
        layout.addWidget(self.log)
        self.setLayout(layout)

    def log_msg(self, msg):
        self.log.append(msg)

    def run_compare(self):
        f1, f2, outdir = self.file1.text(), self.file2.text(), self.output_dir.text()
        if not all([f1, f2, outdir]):
            QMessageBox.warning(self, "缺少文件", "请选择两个 Excel 文件和输出文件夹。")
            return
        try:
            out_path = os.path.join(outdir, "对比结果.xlsx")
            self.log_msg(f"开始对比：\n{f1}\n{f2}")
            highlight_differences_allow_duplicates(f1, f2, output_file=out_path)
            QMessageBox.information(self, "完成", f"结果已保存到：\n{out_path}")
        except Exception as e:
            QMessageBox.critical(self, "错误", str(e))
            self.log_msg(str(e))
